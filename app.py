from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import json
import random
import csv
import io
from openpyxl import load_workbook

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///wheel.db'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

# Database Models
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)

class Name(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)

class WheelConfig(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    locked_result = db.Column(db.String(100), nullable=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin')
@login_required
def admin():
    names = Name.query.all()
    config = WheelConfig.query.first()
    return render_template('admin.html', names=names, config=config)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('admin'))
        
    return render_template('login.html')

@app.route('/api/names', methods=['GET', 'DELETE'])
def get_names():
    if request.method == 'GET':
        names = Name.query.all()
        return jsonify([name.name for name in names])
    elif request.method == 'DELETE':
        name= request.json.get('name')
        Name.query.filter_by(name=name).delete()
        db.session.commit()
        return jsonify({'success': True})

@app.route('/api/spin', methods=['POST'])
def spin():
    config = WheelConfig.query.first()
    if config and config.locked_result:
        result = config.locked_result
        config.locked_result = None
        db.session.commit()
    else:
        names = Name.query.all()
        if names:
            result = random.choice([name.name for name in names])
        else:
            result = None
    return jsonify({'result': result})

@app.route('/api/admin/names', methods=['POST', 'DELETE', 'PUT'])
@login_required
def manage_names():
    if request.method == 'POST':
        name = request.json.get('name')
        db.session.add(Name(name=name))
    elif request.method == 'DELETE':
        name_id = request.json.get('id')
        Name.query.filter_by(id=name_id).delete()
    elif request.method == 'PUT':
        name_id = request.json.get('id')
        new_name = request.json.get('name')
        name = Name.query.get(name_id)
        if name:
            name.name = new_name
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/admin/lock-result', methods=['POST'])
@login_required
def lock_result():
    result = request.json.get('result')
    config = WheelConfig.query.first()
    if not config:
        config = WheelConfig()
        db.session.add(config)
    config.locked_result = result
    db.session.commit()
    return jsonify({'success': True})

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        try:
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower()
            names = []
            
            if file_ext == 'csv':
                # Read CSV file
                stream = io.StringIO(file.stream.read().decode("UTF8"))
                csv_reader = csv.reader(stream)
                for row in csv_reader:
                    if row and row[0].strip():  # Check if row exists and first cell is not empty
                        names.append(row[0].strip())
            else:  # xlsx or xls
                # Read Excel file
                wb = load_workbook(filename=file)
                ws = wb.active
                for row in ws.iter_rows():
                    if row and row[0].value:  # Check if row exists and first cell is not empty
                        names.append(str(row[0].value).strip())
            
            # Add names to database
            added_count = 0
            for name in names:
                if not Name.query.filter_by(name=name).first():
                    new_name = Name(name=name)
                    db.session.add(new_name)
                    added_count += 1
            
            db.session.commit()
            return jsonify({'success': True, 'message': f'Successfully added {added_count} new names'})
            
        except Exception as e:
            return jsonify({'error': str(e)}), 400
    
    return jsonify({'error': 'Invalid file type'}), 400

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Create default admin user if not exists
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                password_hash=generate_password_hash('admin123')
            )
            db.session.add(admin)
            db.session.commit()
    app.run(debug=True)
